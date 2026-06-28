from django.shortcuts import render
from audit.utils import log_action
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Task, Project
from accounts.models import User
from notifications.utils import notify_task_assigned, notify_task_completed
from django.http import JsonResponse
from slack_integration.utils import notify_slack_task_assigned

@login_required
def task_list(request):
    tasks = Task.objects.select_related('assigned_to', 'project').order_by('-created_at')

    # Filter by status
    status = request.GET.get('status')
    if status:
        tasks = tasks.filter(status=status)

    # Filter by priority
    priority = request.GET.get('priority')
    if priority:
        tasks = tasks.filter(priority=priority)

    context = {
        'tasks':        tasks,
        'total':        Task.objects.count(),
        'todo':         Task.objects.filter(status='todo').count(),
        'in_progress':  Task.objects.filter(status='in_progress').count(),
        'done':         Task.objects.filter(status='done').count(),
        'selected_status':   status or '',
        'selected_priority': priority or '',
    }
    return render(request, 'tasks/task_list.html', context)


@login_required
def task_create(request):
    if request.method == 'POST':
        title       = request.POST.get('title')
        description = request.POST.get('description', '')
        priority    = request.POST.get('priority', 'medium')
        status      = request.POST.get('status', 'todo')
        due_date    = request.POST.get('due_date') or None
        assigned_id = request.POST.get('assigned_to') or None
        project_id  = request.POST.get('project') or None

        if not title:
            messages.error(request, 'Task title is required.')
            return redirect('task_create')

        task = Task.objects.create(
            title=title,
            description=description,
            priority=priority,
            status=status,
            due_date=due_date,
            assigned_to_id=assigned_id,
            project_id=project_id,
            created_by=request.user,
        )
        if task.assigned_to_id:
            notify_task_assigned(task, request.user)
            notify_slack_task_assigned(task)
        log_action(request, 'create', 'Task', task.title, task.pk)
        messages.success(request, f'Task "{task.title}" created successfully!')
        return redirect('task_list')

    context = {
        'users':    User.objects.all(),
        'projects': Project.objects.all(),
    }
    return render(request, 'tasks/task_form.html', context)


@login_required
def task_edit(request, pk):
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        task.title       = request.POST.get('title')
        task.description = request.POST.get('description', '')
        task.priority    = request.POST.get('priority', 'medium')
        task.status      = request.POST.get('status', 'todo')
        task.due_date    = request.POST.get('due_date') or None
        assigned_id      = request.POST.get('assigned_to') or None
        project_id       = request.POST.get('project') or None
        task.assigned_to_id = assigned_id
        task.project_id     = project_id
        task.save()
        log_action(request, 'update', 'Task', task.title, task.pk)
        if task.status == 'done':
            notify_task_completed(task)
        notify_task_assigned(task, request.user)
        notify_slack_task_assigned(task)
        messages.success(request, f'Task "{task.title}" updated!')
        return redirect('task_list')

    context = {
        'task':     task,
        'users':    User.objects.all(),
        'projects': Project.objects.all(),
    }
    return render(request, 'tasks/task_form.html', context)


@login_required
def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == 'POST':
        name = task.title
        log_action(request, 'delete', 'Task', task.title, task.pk)
        task.delete()
        messages.success(request, f'Task "{name}" deleted.')
    return redirect('task_list')


@login_required
def task_detail(request, pk):
    task = get_object_or_404(Task, pk=pk)
    return render(request, 'tasks/task_detail.html', {'task': task})


@login_required
def kanban_board(request):
    projects = Project.objects.all()
    selected_project = request.GET.get('project')
    
    tasks_qs = Task.objects.select_related('assigned_to', 'project', 'created_by')
    if selected_project:
        tasks_qs = tasks_qs.filter(project_id=selected_project)
    
    context = {
        'todo':        tasks_qs.filter(status='todo').order_by('-created_at'),
        'in_progress': tasks_qs.filter(status='in_progress').order_by('-created_at'),
        'review':      tasks_qs.filter(status='review').order_by('-created_at'),
        'done':        tasks_qs.filter(status='done').order_by('-created_at'),
        'projects':    projects,
        'selected_project': selected_project or '',
    }
    return render(request, 'tasks/kanban.html', context)


@login_required
def task_update_status(request, pk):
    """AJAX endpoint — update task status from kanban drag"""
    if request.method == 'POST':
        import json
        data   = json.loads(request.body)
        status = data.get('status')
        task   = get_object_or_404(Task, pk=pk)
        valid  = ['todo', 'in_progress', 'review', 'done']
        if status in valid:
            task.status = status
            task.save()
            return JsonResponse({'ok': True, 'status': task.get_status_display()})
    return JsonResponse({'ok': False}, status=400)


@login_required
def add_comment(request, task_pk):
    task = get_object_or_404(Task, pk=task_pk)
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if content:
            from .models import Comment
            Comment.objects.create(
                task=task,
                author=request.user,
                content=content,
            )
            # Send notification to task assignee
            if task.assigned_to and task.assigned_to != request.user:
                try:
                    from notifications.utils import send_notification
                    send_notification(
                        recipient=task.assigned_to,
                        sender=request.user,
                        notif_type='general',
                        title=f'New comment on: {task.title}',
                        message=f'{request.user.get_full_name() or request.user.email} commented: "{content[:80]}"',
                        link=f'/tasks/{task.pk}/',
                    )
                except Exception:
                    pass
            messages.success(request, 'Comment added.')
    return redirect('task_detail', pk=task_pk)


@login_required
def delete_comment(request, pk):
    from .models import Comment
    comment = get_object_or_404(Comment, pk=pk, author=request.user)
    task_pk = comment.task.pk
    if request.method == 'POST':
        comment.delete()
        messages.success(request, 'Comment deleted.')
    return redirect('task_detail', pk=task_pk)

@login_required
def export_tasks_excel(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Header row
    headers = ['ID', 'Title', 'Status', 'Priority', 'Project', 'Assigned To', 'Due Date', 'Created']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='4F46E5')
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')

    # Data rows
    tasks = Task.objects.select_related('project', 'assigned_to').order_by('-created_at')
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.pk)
        ws.cell(row=row, column=2, value=task.title)
        ws.cell(row=row, column=3, value=task.get_status_display())
        ws.cell(row=row, column=4, value=task.get_priority_display())
        ws.cell(row=row, column=5, value=task.project.name if task.project else '')
        ws.cell(row=row, column=6, value=task.assigned_to.get_full_name() if task.assigned_to else 'Unassigned')
        ws.cell(row=row, column=7, value=str(task.due_date) if task.due_date else '')
        ws.cell(row=row, column=8, value=str(task.created_at.date()))

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="nexaops_tasks.xlsx"'
    wb.save(response)
    return response


@login_required
def export_tasks_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from django.http import HttpResponse
    import io

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm, topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#4F46E5'), spaceAfter=6)
    elements.append(Paragraph('NexaOps — Task Report', title_style))
    elements.append(Paragraph(f'Generated on {__import__("datetime").date.today().strftime("%B %d, %Y")}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))

    # Table data
    data = [['#', 'Task Title', 'Status', 'Priority', 'Project', 'Assigned To', 'Due Date']]
    tasks = Task.objects.select_related('project', 'assigned_to').order_by('-created_at')[:100]
    for task in tasks:
        data.append([
            str(task.pk),
            task.title[:50] + ('...' if len(task.title) > 50 else ''),
            task.get_status_display(),
            task.get_priority_display(),
            task.project.name[:20] if task.project else '—',
            task.assigned_to.get_full_name()[:20] if task.assigned_to else 'Unassigned',
            str(task.due_date) if task.due_date else '—',
        ])

    table = Table(data, colWidths=[1.2*cm, 7*cm, 3*cm, 3*cm, 4*cm, 4*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 9),
        ('FONTSIZE',   (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FF')]),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#E2E8F0')),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="nexaops_tasks.pdf"'
    return response